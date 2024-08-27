from fastapi import FastAPI, File, UploadFile, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import FileResponse
import pytesseract
import layoutparser as lp
import cv2
import google.generativeai as genai
from pdf2image import convert_from_path
import numpy as np
from PIL import Image
import pandas as pd
import re
import tempfile
from io import BytesIO
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Configure Google Generative AI API Key
GOOGLE_API_KEY = 'AIzaSyCOY1MkhyEDsOYQC0LEOnQzCCQjEgJhBYI'
genai.configure(api_key=GOOGLE_API_KEY)

# Configure the Tesseract executable path
pytesseract.pytesseract.tesseract_cmd = r'/opt/homebrew/bin/tesseract'

# Initialize FastAPI app
app = FastAPI()

# Mount static files (for serving CSS and JS)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Initialize Jinja2Templates
templates = Jinja2Templates(directory="templates")

# Helper function to clean the text
def clean_text(text):
    text = re.sub(r'\*+', '', text)  # Remove asterisks
    text = re.sub(r'^\d+\.\s*', '', text)  # Remove leading numbers
    text = re.sub(r'-+', '', text)  # Remove hyphens
    return text.strip()

# Process image for OCR and content extraction
def process_image(image: np.array):
    gray = cv2.cvtColor(np.array(image), cv2.COLOR_BGR2GRAY)
    _, img_bin = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

    # Perform OCR using Tesseract
    ocr_result = pytesseract.image_to_string(img_bin)

    # Use Google Generative AI to extract invoice details
    model = genai.GenerativeModel('gemini-pro')

    prompt = f"""
    Extract the following relevant details from the invoice text:
    1. Invoice Number
    2. Invoice Date
    3. Buyer Name
    4. Buyer Address
    5. Buyer GSTIN
    6. Seller Name
    7. Seller Address
    8. Seller GSTIN
    9. State Name
    10. PO Number
    11. Invoice Amount
    12. Taxable Amount
    13. Total Amount
    14. Tax Amount (CGST, SGST, IGST)
    15. Total Tax Amount
    16. TCS Amount
    17. IRN Number
    18. Receiver GSTIN
    19. Receiver Name
    20. Vendor GSTIN
    21. Vendor Name
    22. Vendor Code
    23. Remarks
    24. Other relevant financial details

    The invoice text is as follows:
    {ocr_result}
    """

    response = model.generate_content(prompt)
    structured_output = response.text

    # Parse structured output
    details = []
    for line in structured_output.split("\n"):
        if ":" in line:
            key, value = line.split(":", 1)
            key = clean_text(key)
            value = clean_text(value)
            details.append([key, value])

    if details:
        df = pd.DataFrame(details, columns=["Field", "Value"])
        return df
    return None

# Root endpoint for the website
@app.get("/")
async def render_form(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

# Endpoint to handle form submission and display extracted invoice details
@app.post("/upload-invoice/")
async def process_invoice(request: Request, file: UploadFile = File(...)):
    try:
        contents = await file.read()

        if file.content_type == 'application/pdf':
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                tmp_file.write(contents)
                pdf_path = tmp_file.name

            # Convert PDF to image
            images = convert_from_path(pdf_path)
            df = process_image(images[0])

        elif file.content_type.startswith('image/'):
            image = Image.open(BytesIO(contents))
            df = process_image(image)

        if df is not None:
            # Save DataFrame to Excel
            output_excel_path = os.path.join("static", "extracted_invoice_details.xlsx")

            # Create a new Excel file with the format similar to the one you provided
            wb = Workbook()
            ws = wb.active

            # Add the fields as column headers
            for col_idx, col_name in enumerate(df['Field'], 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)  # Make the headers bold

            # Add the extracted values as the first row of data
            for col_idx, value in enumerate(df['Value'], 1):
                ws.cell(row=2, column=col_idx, value=value)

            # Save the Excel file
            wb.save(output_excel_path)

            # Render the result on the page and provide the download link
            result_html = df.to_html(index=False, classes="table table-striped")
            return templates.TemplateResponse("index.html", {"request": request, "result": result_html, "excel_link": f"/static/extracted_invoice_details.xlsx"})
        else:
            return templates.TemplateResponse("index.html", {"request": request, "error": "Failed to extract details"})

    except Exception as e:
        return templates.TemplateResponse("index.html", {"request": request, "error": str(e)})
